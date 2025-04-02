import pandas as pd
from typing import Optional
import logging
import asyncio
import aiohttp
import openpyxl

logger = logging.getLogger(__name__)

class WbYMProcessor:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self._df_wb: Optional[pd.DataFrame] = None
        self._df_ym: Optional[pd.DataFrame] = None

        # Переменные для OpenPyXL
        self._wb = None         # Workbook
        self._ws_wb = None      # Лист WB
        self._ym_id_col_idx = None  # Номер колонки (целое число) для YM_id

    def read_wb_data(self, sheet_name: str = "WB") -> Optional[pd.DataFrame]:
        """
        Читает лист 'WB' через pandas, проверяет нужные колонки,
        а также открывает Excel-файл через openpyxl,
        чтобы потом можно было обновлять столбец YM_id «на лету».
        """
        required_columns = ['parent_id', 'parent_name', 'subject_id', 'subject_name', 'YM_id']
        try:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name)
            logger.info(f"Файл '{self.file_path}' успешно прочитан (лист '{sheet_name}').")
            logger.debug(f"Прочитано строк: {df.shape[0]}, столбцов: {df.shape[1]}")
        except Exception as e:
            logger.error(f"Ошибка чтения файла '{self.file_path}' (лист '{sheet_name}'): {e}")
            return None

        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            logger.error(
                f"В листе '{sheet_name}' отсутствуют колонки: {', '.join(missing_columns)}"
            )
            return None

        # Сохраняем DataFrame в памяти
        self._df_wb = df[required_columns].copy()
        logger.info(f"WB-данные успешно сохранены. Итоговая форма: {self._df_wb.shape}")

        # Загружаем Workbook через openpyxl, чтобы иметь доступ к ячейкам
        self._wb = openpyxl.load_workbook(self.file_path)
        if sheet_name not in self._wb.sheetnames:
            logger.error(f"В файле нет листа '{sheet_name}'")
            return self._df_wb

        self._ws_wb = self._wb[sheet_name]

        # Находим индекс колонки "YM_id" в файле (по имени столбца в первой строке)
        # Предполагается, что в строке 1 – заголовки.
        header_row = next(self._ws_wb.iter_rows(min_row=1, max_row=1, values_only=True))
        # header_row – это кортеж с названиями колонок, например ('parent_id', 'parent_name', ...)
        # Нужно найти индекс "YM_id".
        try:
            self._ym_id_col_idx = header_row.index("YM_id") + 1  # openpyxl колонки нумеруются с 1
            logger.debug(f"Колонка 'YM_id' найдена в Excel: индекс={self._ym_id_col_idx}")
        except ValueError:
            logger.error(f"Не смогли найти колонку 'YM_id' в Excel. Проверьте заголовки.")
            self._ym_id_col_idx = None

        return self._df_wb

    def read_ym_data(self, sheet_name: str = "YM") -> Optional[pd.DataFrame]:
        """
        Читает лист 'YM' и сохраняет в self._df_ym.
        """
        required_columns = ['last_id', 'last_name']
        try:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name)
            logger.info(f"Файл '{self.file_path}' успешно прочитан (лист '{sheet_name}').")
            logger.debug(f"Прочитано строк: {df.shape[0]}, столбцов: {df.shape[1]}")
        except Exception as e:
            logger.error(f"Ошибка чтения файла '{self.file_path}' (лист '{sheet_name}'): {e}")
            return None

        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            logger.error(
                f"В листе '{sheet_name}' отсутствуют колонки: {', '.join(missing_columns)}"
            )
            return None

        self._df_ym = df[required_columns].copy()
        logger.info(f"YM-данные успешно сохранены. Итоговая форма: {self._df_ym.shape}")
        return self._df_ym

    import aiohttp
    import logging

    logger = logging.getLogger(__name__)

    async def fetch_subject_id(self, session: aiohttp.ClientSession, name_product: str) -> Optional[int]:
        """
        Асинхронно делает запрос к API Wildberries, чтобы получить subject_id.
        Если subject_id отсутствует, возвращаем subjectParentId.
        """
        url = (
            "https://search.wb.ru/exactmatch/ru/common/v9/"
            f"search?appType=1&curr=page=1&query={name_product}&resultset=catalog"
        )
        logger.debug(f"[fetch_subject_id] Запрос к WB API: {url}")

        try:
            async with session.get(url) as response:
                # Проверяем статус ответа
                response.raise_for_status()

                # Проверяем Content-Type, прежде чем парсить
                content_type = response.headers.get("Content-Type", "").lower()
                if "application/json" not in content_type:
                    # Если ответ не в JSON-формате, логируем и завершаем
                    text_body = await response.text()
                    logger.error(
                        f"WB вернул не JSON (content-type={content_type}). "
                        f"Ответ (200 символов): {text_body[:200]}"
                    )
                    return None

                # Пробуем распарсить JSON
                try:
                    data = await response.json()
                except aiohttp.ContentTypeError:
                    text_body = await response.text()
                    logger.error(
                        f"Ошибка парсинга JSON для '{name_product}'. "
                        f"Текст ответа (200 символов): {text_body[:200]}"
                    )
                    return None

        except aiohttp.ClientResponseError as e:
            # Ошибка сервера: e.status, e.message и т.д.
            logger.error(
                f"Ошибка при запросе к API Wildberries для '{name_product}': "
                f"status={e.status}, message={e.message}, url='{e.request_info.real_url}'"
            )
            return None
        except aiohttp.ClientConnectionError as e:
            # Ошибка сети (нет доступа, таймаут и т.п.)
            logger.error(f"Ошибка сети при запросе к API Wildberries для '{name_product}': {e}")
            return None
        except Exception as e:
            # Любая другая неожиданная ошибка
            logger.error(f"Ошибка при запросе к API Wildberries для '{name_product}': {e}")
            return None

        # Обрабатываем данные
        products = data.get("data", {}).get("products", [])
        if not products:
            logger.debug(f"[fetch_subject_id] Пустой список products для '{name_product}'.")
            return None

        # Берём первый товар из списка
        product = products[0]
        subject_id = product.get("subjectId")
        if subject_id is not None:
            logger.debug(f"[fetch_subject_id] Для '{name_product}' получен subjectId: {subject_id}")
            return subject_id
        else:
            parent_id = product.get("subjectParentId")
            logger.debug(
                f"[fetch_subject_id] Для '{name_product}' subjectId отсутствует. "
                f"Возвращаем parentId: {parent_id}"
            )
            return parent_id

    async def _process_row(self, session: aiohttp.ClientSession, idx: int, last_name: str, last_id: int):
        """
        Вспомогательный метод-корутина: для одной строки YM:
        1) async fetch_subject_id
        2) вернуть (idx, subject_id, last_id)
        """
        subject_id = await self.fetch_subject_id(session, last_name)
        return idx, subject_id, last_id

    async def async_update_wb_with_ym(self) -> None:
        """
        Асинхронно проходит по строкам листа YM:
        - Для каждого last_name параллельно (через asyncio) получает subject_id;
        - Ищет в WB строку с таким subject_id;
        - И «вживую» (через openpyxl) записывает last_id в колонку YM_id Excel-файла,
          сразу же вызывая save(), чтобы не потерять изменения при прерывании.
        """
        if self._df_wb is None or self._df_ym is None:
            logger.error("Сначала нужно считать данные из обоих листов (WB и YM).")
            return
        if not self._ws_wb or self._ym_id_col_idx is None:
            logger.error("Невозможно обновлять Excel: нет доступной worksheet или YM_id-колонки.")
            return

        tasks = []
        async with aiohttp.ClientSession() as session:
            for idx, row in self._df_ym.iterrows():
                last_name = row['last_name']
                last_id = row['last_id']
                tasks.append(self._process_row(session, idx, last_name, last_id))

            results = await asyncio.gather(*tasks, return_exceptions=True)

        updated_count = 0

        # Сопоставляем результаты с DataFrame WB и сразу пишем в Excel
        for result in results:
            if isinstance(result, Exception):
                logger.error(f"Исключение во время обработки строки: {result}")
                continue

            idx, subject_id, last_id = result
            if subject_id is None:
                logger.warning(
                    f"[{idx}] Не найден subject_id (API вернул пустой результат) для last_name='{self._df_ym.at[idx, 'last_name']}'."
                )
                continue

            # Ищем строки в _df_wb, у которых совпадает subject_id
            mask = self._df_wb['subject_id'] == subject_id
            if not mask.any():
                logger.warning(
                    f"[{idx}] subject_id={subject_id} не найден в листе WB."
                )
                continue

            # Обновляем в DataFrame
            self._df_wb.loc[mask, 'YM_id'] = last_id

            # Обновляем в Excel «по месту»:
            # Так как DataFrame и Excel (openpyxl) имеют одну и ту же логику индексирования строк —
            # мы считаем, что в Excel первая строка это заголовки, значит данные начинаются со строки 2
            df_rows_to_update = self._df_wb[mask].index
            for df_row in df_rows_to_update:
                excel_row = df_row + 2  # смещение на заголовок
                self._ws_wb.cell(row=excel_row, column=self._ym_id_col_idx, value=last_id)
                updated_count += 1

            # После каждой записи – сохраняем файл.
            # Это медленно для больших файлов, но гарантирует, что не потеряем изменения при прерывании
            self._wb.save(self.file_path)
            logger.info(
                f"[{idx}] subject_id={subject_id} → YM_id={last_id}. Обновили {len(df_rows_to_update)} стр. в Excel."
            )

        logger.info(f"[async_update_wb_with_ym] Итог: обновлено строк: {updated_count}.")

    @property
    def wb_data(self) -> Optional[pd.DataFrame]:
        """Возвращает текущую копию данных из листа WB (в памяти)."""
        return self._df_wb

    @property
    def ym_data(self) -> Optional[pd.DataFrame]:
        """Возвращает текущую копию данных из листа YM (в памяти)."""
        return self._df_ym
