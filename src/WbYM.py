import pandas as pd
from typing import Optional
import logging
import requests

# Добавляем openpyxl, чтобы сохранять структуру Excel и обновлять только нужные ячейки
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

class WbYMProcessor:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self._df_wb: Optional[pd.DataFrame] = None
        self._df_ym: Optional[pd.DataFrame] = None

    def read_wb_data(self, sheet_name: str = "WB") -> Optional[pd.DataFrame]:
        """Читает лист 'WB' и сохраняет в self._df_wb."""
        required_columns = ['parent_id', 'parent_name', 'subject_id', 'subject_name', 'YM_id']
        try:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name)
            logger.info(f"Файл '{self.file_path}' успешно прочитан (лист '{sheet_name}').")
            logger.debug(f"Прочитано строк: {df.shape[0]}, столбцов: {df.shape[1]}")
        except Exception as e:
            logger.error(f"Ошибка чтения файла '{self.file_path}' (лист '{sheet_name}'): {e}")
            return None

        # Проверяем наличие нужных колонок
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            logger.error(
                f"В листе '{sheet_name}' отсутствуют колонки: {', '.join(missing_columns)}"
            )
            return None

        self._df_wb = df[required_columns].copy()
        logger.info(f"WB-данные успешно сохранены. Итоговая форма: {self._df_wb.shape}")
        return self._df_wb

    def read_ym_data(self, sheet_name: str = "YM") -> Optional[pd.DataFrame]:
        """Читает лист 'YM' и сохраняет в self._df_ym."""
        required_columns = ['last_id', 'last_name']
        try:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name)
            logger.info(f"Файл '{self.file_path}' успешно прочитан (лист '{sheet_name}').")
            logger.debug(f"Прочитано строк: {df.shape[0]}, столбцов: {df.shape[1]}")
        except Exception as e:
            logger.error(f"Ошибка чтения файла '{self.file_path}' (лист '{sheet_name}'): {e}")
            return None

        # Аналогичная проверка наличия необходимых колонок
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            logger.error(
                f"В листе '{sheet_name}' отсутствуют колонки: {', '.join(missing_columns)}"
            )
            return None

        self._df_ym = df[required_columns].copy()
        logger.info(f"YM-данные успешно сохранены. Итоговая форма: {self._df_ym.shape}")
        return self._df_ym

    def get_subject_id(self, name_product: str) -> Optional[int]:
        """
        Делает запрос к API Wildberries, чтобы получить subject_id.
        Если subject_id отсутствует, возвращаем subjectParentId.
        """
        url = (
            "https://search.wb.ru/exactmatch/ru/common/v9/"
            f"search?appType=1&curr=page=1&query={name_product}&resultset=catalog"
        )
        logger.debug(f"Запрос к WB API: {url}")
        try:
            response = requests.get(url)
            response.raise_for_status()
            data = response.json()
        except requests.RequestException as e:
            logger.error(f"Ошибка при запросе к API Wildberries для '{name_product}': {e}")
            return None

        products = data.get("data", {}).get("products", [])
        if not products:
            logger.debug(f"API WB вернул пустой список products для '{name_product}'.")
            return None

        product = products[0]
        subject_id = product.get("subjectId")
        if subject_id is not None:
            logger.debug(f"Для '{name_product}' получен subjectId: {subject_id}")
            return subject_id
        else:
            parent_id = product.get("subjectParentId")
            logger.debug(f"Для '{name_product}' subjectId отсутствует, возвращаем parentId: {parent_id}")
            return parent_id

    def update_wb_with_ym(self, chunk_size: int = 100) -> None:
        """
        Проходим по строкам листа YM батчами по chunk_size:
        - Для каждого last_name получаем subject_id из API;
        - Ищем в WB строки с таким subject_id;
        - Проставляем в их YM_id значение last_id из YM;
        - После обработки каждого батча (100 записей), сохраняем обновлённый YM_id обратно в Excel,
          не ломая структуру таблицы, обновляем только колонку 'YM_id'.
        """
        if self._df_wb is None or self._df_ym is None:
            logger.error("Сначала нужно считать данные из обоих листов (WB и YM).")
            return

        updated_count = 0
        total_rows = len(self._df_ym)

        # Делим self._df_ym на куски по chunk_size
        for start in range(0, total_rows, chunk_size):
            end = start + chunk_size
            chunk = self._df_ym.iloc[start:end]

            # Обновляем данные в self._df_wb (только в памяти)
            for idx, row in chunk.iterrows():
                last_name = row['last_name']
                last_id = row['last_id']

                subject_id = self.get_subject_id(last_name)
                if subject_id is not None:
                    mask = self._df_wb['subject_id'] == subject_id
                    if mask.any():
                        self._df_wb.loc[mask, 'YM_id'] = last_id
                        logger.info(
                            f"[{idx}] Для last_name='{last_name}' найден subject_id={subject_id}. "
                            f"Установлен YM_id={last_id}"
                        )
                        updated_count += mask.sum()
                    else:
                        logger.warning(
                            f"[{idx}] subject_id={subject_id} из API не найден в листе WB."
                        )
                else:
                    logger.warning(
                        f"[{idx}] Не найден subject_id для last_name='{last_name}' (API вернул пустой результат)."
                    )

            # Каждый обработанный батч (100 записей) сразу сохраняем в Excel
            self._save_updated_ym_id_to_excel()
            logger.info(f"Обработано записей: {end if end < total_rows else total_rows} / {total_rows}")

        logger.info(f"update_wb_with_ym завершён. Всего обновлено строк: {updated_count}.")

    def _save_updated_ym_id_to_excel(self, sheet_name: str = "WB") -> None:
        """
        Сохраняет текущие значения YM_id из self._df_wb обратно в Excel,
        не трогая структуру листа и другие столбцы.
        """
        if self._df_wb is None:
            logger.error("Нет данных в self._df_wb для сохранения.")
            return

        try:
            # Загружаем существующую книгу
            wb = load_workbook(self.file_path)
            ws = wb[sheet_name]

            # Считываем заголовки, чтобы найти индекс столбца YM_id
            headers = [cell.value for cell in ws[1]]
            if 'YM_id' not in headers:
                logger.error(f"Не найден столбец 'YM_id' в листе '{sheet_name}'.")
                return
            ym_id_col_idx = headers.index('YM_id') + 1  # +1 из-за 1-базовой индексации в Excel

            # Перебираем строки DataFrame и обновляем только YM_id
            for i, row_data in self._df_wb.iterrows():
                # В Excel строки начинаются с 1, первая строка — заголовки
                excel_row = i + 2
                ws.cell(row=excel_row, column=ym_id_col_idx).value = row_data['YM_id']

            wb.save(self.file_path)
            logger.info(f"Данные YM_id успешно сохранены в лист '{sheet_name}'.")
        except Exception as e:
            logger.error(f"Ошибка при сохранении YM_id в Excel: {e}")

    @property
    def wb_data(self) -> Optional[pd.DataFrame]:
        """Возвращает текущую копию данных из листа WB."""
        return self._df_wb

    @property
    def ym_data(self) -> Optional[pd.DataFrame]:
        """Возвращает текущую копию данных из листа YM."""
        return self._df_ym
