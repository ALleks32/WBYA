from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from src.logger import logger
from src.data_readers import WbYMReader, CabinetTableReader

class YMIdFiller:
    """
    Класс для заполнения колонки YM_id в файле wb-ym.xlsx на основе данных
    из таблиц кабинетов.

    Алгоритм:
      1. Берем значение из поля subject_name и ищем его в колонке Категория_YMname.
      2. Если совпадение найдено – в YM_id записываем соответствующий Категория_YMid.
      3. Если по subject_name совпадения нет, пробуем по parent_name.
    """

    def __init__(self, wbym_df: pd.DataFrame, cabinet_df: pd.DataFrame):
        self.wbym_df = wbym_df
        self.cabinet_df = cabinet_df

    def fill_yamid(self) -> None:
        # Формируем словарь сопоставления: Категория_YMname -> Категория_YMid
        mapping = dict(zip(self.cabinet_df['Категория_YMname'], self.cabinet_df['Категория_YMid']))

        for idx, row in self.wbym_df.iterrows():
            ym_value = None
            subject_name = row.get('subject_name')
            parent_name = row.get('parent_name')

            # Сначала пробуем найти совпадение по subject_name
            if subject_name in mapping:
                ym_value = mapping[subject_name]
            # Если не найдено, ищем по parent_name
            elif parent_name in mapping:
                ym_value = mapping[parent_name]

            if ym_value:
                self.wbym_df.at[idx, 'YM_id'] = ym_value
                logger.info(
                    f"Строке {idx} присвоен YM_id '{ym_value}' (по {'subject_name' if subject_name in mapping else 'parent_name'}).")
            else:
                logger.warning(
                    f"Для строки {idx} не найдено совпадение по subject_name='{subject_name}' и parent_name='{parent_name}'.")


def filler_to_name_category():
    # Задаем пути к файлам и папкам
    wbym_file = Path("data") / "wb-ym.xlsx"
    cabinets_folder = Path("data") / "cabinets"

    # Чтение файла wb-ym.xlsx с листа 'WB'
    wbym_reader = WbYMReader(str(wbym_file))
    wbym_df = wbym_reader.read_data(sheet_name="WB")
    if wbym_df is None:
        logger.error("Не удалось прочитать файл wb-ym.xlsx, лист 'WB'.")
        return

    # Чтение и объединение файлов кабинетов из папки data/cabinets
    cabinet_reader = CabinetTableReader(str(cabinets_folder))
    cabinet_df = cabinet_reader.read_data()
    if cabinet_df is None:
        logger.error("Не удалось прочитать файлы кабинетов из папки cabinets.")
        return

    # Заполнение колонки YM_id в DataFrame
    filler = YMIdFiller(wbym_df, cabinet_df)
    filler.fill_yamid()

    # Обновляем только значения в столбце YM_id в файле wb-ym.xlsx, лист 'WB'
    try:
        wb = load_workbook(filename=str(wbym_file))
        ws = wb["WB"]

        # Определяем номер столбца для 'YM_id' по заголовку в первой строке
        ym_id_col = None
        for cell in ws[1]:
            if cell.value == "YM_id":
                ym_id_col = cell.column  # в openpyxl 2.6+ возвращается число
                break

        if ym_id_col is None:
            logger.error("В листе 'WB' не найден столбец 'YM_id'.")
            return

        # Обновляем значения столбца YM_id в каждой строке, соответствующей DataFrame
        # Предполагается, что заголовок находится в первой строке, а данные начинаются со второй.
        for idx, row in wbym_df.iterrows():
            excel_row = idx + 2  # индекс 0 соответствует строке 2 в Excel
            new_value = row["YM_id"]
            ws.cell(row=excel_row, column=ym_id_col, value=new_value)

        wb.save(str(wbym_file))
        logger.info(f"Столбец 'YM_id' на листе 'WB' успешно обновлен в файле: {wbym_file}")
    except Exception as e:
        logger.error(f"Ошибка при обновлении книги: {e}")

