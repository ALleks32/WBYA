import pandas as pd
from typing import Optional
from src.logger import logger
import requests
from openpyxl import load_workbook
from src.wb_api import get_subject_id

class WbYMProcessor:
    def __init__(self, file_path: str, df_wb: Optional[pd.DataFrame], df_ym: Optional[pd.DataFrame]):
        self.file_path = file_path
        self._df_wb: Optional[pd.DataFrame] = df_wb
        self._df_ym: Optional[pd.DataFrame] = df_ym


    def update_wb_with_ym(self, chunk_size: int = 100) -> None:
        """
        Проходим по строкам листа YM батчами по chunk_size:
        - Для каждого last_name получаем subject_id из API;
        - Ищем в WB строки с таким subject_id;
        - Проставляем в их YM_id значение last_id из YM;
        - Одновременно создаём/обновляем в WB и колонку YM_name тем же last_name.
        - После обработки каждого батча (100 записей), сохраняем обновления обратно в Excel,
          не ломая структуру таблицы (обновляем только нужные колонки).
        """
        if self._df_wb is None or self._df_ym is None:
            logger.error("Сначала нужно считать данные из обоих листов (WB и YM).")
            return

        # Если в DF WB ещё нет колонки YM_name, создадим пустую
        if 'YM_name' not in self._df_wb.columns:
            self._df_wb['YM_name'] = None

        updated_count = 0
        total_rows = len(self._df_ym)

        # Делим self._df_ym на куски по chunk_size
        for start in range(0, total_rows, chunk_size):
            end = start + chunk_size
            chunk = self._df_ym.iloc[start:end]

            # Обновляем данные в self._df_wb (только в памяти)
            for idx, row in chunk.iterrows():
                last_name = row['last_name']
                last_id = row['last_id']

                subject_id = get_subject_id(last_name)
                if subject_id is not None:
                    mask = self._df_wb['subject_id'] == subject_id
                    if mask.any():
                        self._df_wb.loc[mask, 'YM_id'] = last_id
                        self._df_wb.loc[mask, 'YM_name'] = last_name  # <-- Записываем название категории
                        logger.info(
                            f"[{idx}] Для last_name='{last_name}' найден subject_id={subject_id}. "
                            f"Установлен YM_id={last_id}, YM_name='{last_name}'"
                        )
                        updated_count += mask.sum()
                    else:
                        logger.warning(
                            f"[{idx}] subject_id={subject_id} из API не найден в листе WB."
                        )
                else:
                    logger.warning(
                        f"[{idx}] Не найден subject_id для last_name='{last_name}' (API вернул пустой результат)."
                    )

            # Каждый обработанный батч сразу сохраняем в Excel
            self._save_updated_ym_data_to_excel()
            logger.info(f"Обработано записей: {end if end < total_rows else total_rows} / {total_rows}")

        logger.info(f"update_wb_with_ym завершён. Всего обновлено строк: {updated_count}.")

    def _save_updated_ym_data_to_excel(self, sheet_name: str = "WB") -> None:
        """
        Сохраняет текущие значения YM_id и YM_name из self._df_wb обратно в Excel,
        не трогая структуру листа и другие столбцы.
        Если столбца YM_name нет в самом Excel, добавляем его рядом с YM_id.
        """
        if self._df_wb is None:
            logger.error("Нет данных в self._df_wb для сохранения.")
            return

        try:
            wb = load_workbook(self.file_path)
            ws = wb[sheet_name]

            # Читаем заголовки
            headers = [cell.value for cell in ws[1]]
            # Проверяем наличие YM_id
            if 'YM_id' not in headers:
                logger.error(f"Не найден столбец 'YM_id' в листе '{sheet_name}'.")
                return

            ym_id_col_idx = headers.index('YM_id') + 1

            # Проверяем наличие YM_name
            if 'YM_name' not in headers:
                # Вставляем новый столбец сразу после 'YM_id'
                new_col_index = ym_id_col_idx + 1
                ws.insert_cols(new_col_index)
                # Подпишем заголовок новой колонки
                ws.cell(row=1, column=new_col_index, value='YM_name')

                # Нужно обновить headers, иначе ниже будем искать по старому списку
                headers.insert(new_col_index - 1, 'YM_name')

            ym_name_col_idx = headers.index('YM_name') + 1

            # Теперь проставляем значения YM_id и YM_name по строкам
            for i, row_data in self._df_wb.iterrows():
                # В Excel строки начинаются с 1, первая строка — заголовки
                excel_row = i + 2

                # Обновляем YM_id
                ws.cell(row=excel_row, column=ym_id_col_idx).value = row_data['YM_id']
                # Обновляем YM_name
                ws.cell(row=excel_row, column=ym_name_col_idx).value = row_data['YM_name']

            wb.save(self.file_path)
            logger.info(f"Данные YM_id и YM_name успешно сохранены в лист '{sheet_name}'.")
        except Exception as e:
            logger.error(f"Ошибка при сохранении YM_id/YM_name в Excel: {e}")

